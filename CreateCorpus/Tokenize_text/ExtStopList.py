# -*- coding: utf-8 -*-
"""
Created on Tue Dec 21 22:13:34 2021

@author: PC
"""
def ExtStopWords():
    import openpyxl
    import nltk
    from nltk.corpus import stopwords 
    wb = openpyxl.load_workbook('./CreateCorpus/Tokenize_text/AidaStopWords.xlsx')
    sheets = wb.sheetnames
    indSheet = 0
    stopwords = nltk.corpus.stopwords.words('english')
    for sheet in sheets:
        wb.active = indSheet
        sheet = wb.active
        indCell = 0
        while indCell < 1000:
            indCell = indCell + 1
            cell = sheet.cell(row = indCell, column = 1)
            if cell.value == None:
                break
            t = ''; s = cell.value; i = 0
            while i < len(s):
                if s[i] >= 'a' and s[i] <= 'z' or s[i] >= 'A' and s[i] <= 'Z' or s[i] == ' ': 
                    t = t + s[i]
                i = i + 1
            t = t.lower()
            stopwords.append(str(t).strip())
        indSheet = indSheet + 1
    stopWords = stopwords
    return(stopWords)
